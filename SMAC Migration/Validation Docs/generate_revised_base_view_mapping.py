"""Generate revised_base_view SAC vs SMAC column mapping (Excel + Markdown)."""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = r"SMAC Migration\Validation Docs\revised_base_view_column_mapping.xlsx"
OUTPUT_MD = r"SMAC Migration\Validation Docs\revised_base_view_column_mapping.md"


def md_escape(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ")


def dataframe_to_md_table(df: pd.DataFrame) -> str:
    headers = [md_escape(c) for c in df.columns]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for _, row in df.iterrows():
        lines.append("| " + " | ".join(md_escape(row[c]) for c in df.columns) + " |")
    return "\n".join(lines)

rows = [
    # Seafarer identity
    ("SEAFARER_ID", "SEAFARER_ID", "CAST(B.ID AS INT)", "landing_zone.db_sac_prod_seafarer_public.seafarers (B)", "B.id", "curated_db.db_smac_prod_navitasai_crewing_public.seafarers (B)", "Direct ID; SMAC uses UUID string instead of INT", "High", "Data type change INT→UUID. Join/reporting must use CREW_CODE for cross-system validation."),
    ("FIRST_NAME", "FIRST_NAME", "B.FIRST_NAME", "seafarers (B)", "B.first_name", "seafarers (B)", "Direct column mapping", "High", ""),
    ("MIDDLE_NAME", "MIDDLE_NAME", "B.MIDDLE_NAME", "seafarers (B)", "B.middle_name", "seafarers (B)", "Direct column mapping", "High", ""),
    ("LAST_NAME", "LAST_NAME", "B.LAST_NAME", "seafarers (B)", "B.last_name", "seafarers (B)", "Direct column mapping", "High", ""),
    ("SEAFARER_NAME", "SEAFARER_NAME", "CONCAT(COALESCE(B.FIRST_NAME,' '), ' ', COALESCE(B.MIDDLE_NAME,' '), ' ', COALESCE(B.LAST_NAME,' '))", "seafarers (B)", "CONCAT(COALESCE(B.first_name,' '),' ',COALESCE(B.middle_name,' '),' ',COALESCE(B.last_name,' '))", "seafarers (B)", "Same concatenation logic", "High", "Extra spaces possible when middle name null (same in both)."),
    ("USER_ID", "USER_ID", "B.UUID", "seafarers (B)", "B.id", "seafarers (B)", "SAC used UUID column; SMAC uses seafarer id (UUID)", "High", "SMAC B.id replaces SAC B.UUID — same purpose for profile links."),
    ("AHOY_STATUS", "AHOY_STATUS", "CASE WHEN B.IDENTITY_USER_ID IS NOT NULL THEN 'Ahoy Installed' ELSE 'Ahoy Not Installed' END", "seafarers (B)", "CASE WHEN B.identity_profile_id IS NOT NULL THEN 'Ahoy Installed' ELSE 'Ahoy Not Installed' END", "seafarers (B)", "Same CASE logic; column renamed identity_user_id→identity_profile_id", "High", ""),
    ("CREW_CODE", "CREW_CODE", "B.CREW_CODE", "seafarers (B)", "B.crew_code", "seafarers (B)", "Direct column mapping", "High", "Primary cross-system join key."),
    ("OLD_CREW_CODE", "OLD_CREW_CODE", "B.OLD_CREW_CODE", "seafarers (B)", "B.old_crew_code", "seafarers (B)", "Direct column mapping", "High", ""),
    ("CURRENT_STATUS", "CURRENT_STATUS", "CASE mapping B.STATE values (sign_on→Sign On, etc.)", "seafarers (B)", "PS.name", "profile_states (PS) via B.profile_state_id = PS.id", "SAC maps raw state enum; SMAC uses profile_states lookup name", "Medium", "SMAC returns master name (e.g. 'Sign On'); verify PS.name matches SAC CASE output exactly."),
    ("SEAFARER_TYPE", "SEAFARER_TYPE", "CASE WHEN UPPER(B.STATE) IN ('REGISTERED','APPLIED','SELECTED') THEN 'External Seafarers' ELSE 'Internal Seafarers' END", "seafarers (B)", "CASE WHEN UPPER(PS.code) IN ('REGISTERED','APPLIED','SELECTED') THEN 'External Seafarers' ELSE 'Internal Seafarers' END", "seafarers (B) + profile_states (PS)", "Same classification; SMAC uses PS.code instead of B.STATE", "High", ""),
    ("ANNIVERSARY_DATE", "ANNIVERSARY_DATE", "B.ANNIVERSARY_DATE", "seafarers (B)", "SP.anniversary_date", "seafarer_profile (SP) via SP.seafarer_id = B.id", "Moved from seafarers to seafarer_profile in SMAC", "High", "JOIN: LEFT JOIN seafarer_profile SP ON B.id = SP.seafarer_id AND SP.deleted_at IS NULL"),
    ("RANK_ID", "RANK_ID", "CAST(B.RANK_ID AS INT)", "seafarers (B)", "B.rank_id", "seafarers (B)", "Direct FK; SMAC UUID vs SAC INT", "High", "Type change INT→UUID."),
    ("GENDER_NAME", "GENDER_NAME", "CASE B.GENDER '1'→Female, '0'→Male, else Unknown", "seafarers (B)", "COALESCE(GEN.name, 'Unknown')", "genders (GEN) via B.gender_id = GEN.id", "SAC inline enum; SMAC lookup table", "High", ""),
    ("DATE_OF_BIRTH", "DATE_OF_BIRTH", "B.DATE_OF_BIRTH", "seafarers (B)", "B.date_of_birth", "seafarers (B)", "Direct column mapping", "High", ""),
    ("AGE", "AGE", "datediff(YEAR, CAST(B.DATE_OF_BIRTH AS DATE), current_date())", "seafarers (B)", "datediff(YEAR, CAST(B.date_of_birth AS DATE), current_date())", "seafarers (B)", "Same age calculation", "High", ""),
    ("AGE_CATEGORY", "AGE_CATEGORY", "CASE AGE <30 / 30-49 / 50+", "Derived", "Same CASE on AGE", "Derived", "Same bucketing logic", "High", ""),
    ("CREATED_AT", "CREATED_AT", "B.CREATED_AT", "seafarers (B)", "B.created_at", "seafarers (B)", "Direct column mapping", "High", ""),
    ("PROFILE_STATUS", "PROFILE_STATUS", "CASE UPPER(B.IS_ACTIVE)='TRUE' THEN 'Active Seafarer' ELSE 'Inactive Seafarer'", "seafarers (B)", "CASE PST.code='ACTIVE' THEN 'Active Seafarer' ELSE 'Inactive Seafarer'", "seafarer_profile_statuses (PST) via B.profile_status_id", "SAC boolean flag; SMAC status lookup", "High", "JOIN: LEFT JOIN seafarer_profile_statuses PST ON B.profile_status_id = PST.id"),
    ("ONBOARD_SAILING_STATUS", "ONBOARD_SAILING_STATUS", "Active+Sign On→Onboard; J.FROM_DATE/TO_DATE both NULL→No Past Records; else Onleave", "seafarers + sea_experiences", "PST.code='ACTIVE' AND PS.code='SIGNON'→Onboard; J.sign_on/off NULL→No Past Records; else Onleave", "seafarers + profile_states + sea_experiences", "Same 3-way logic; SMAC uses PS.code='SIGNON' vs SAC 'Sign On'", "Medium", "SAC checks J.FROM_DATE/TO_DATE; SMAC checks J.sign_on_date/sign_off_date. Should be equivalent."),
    ("AVAILABILITY_DATE", "AVAILABILITY_DATE", "B.AVAILABILITY_DATE", "seafarers (B)", "B.availability_date", "seafarers (B)", "Direct column mapping", "High", ""),
    ("AVAILABILITY_MONTH", "AVAILABILITY_MONTH", "month(CAST(B.AVAILABILITY_DATE AS DATE))", "seafarers (B)", "month(CAST(B.availability_date AS DATE))", "seafarers (B)", "Same extraction", "High", ""),
    ("CDC_NUMBER", "CDC_NUMBER", "B.CDC_NUMBER", "seafarers (B)", "B.cdc_number", "seafarers (B)", "Direct column mapping", "High", ""),
    ("APPRAISAL_LINK", "APPRAISAL_LINK", "CONCAT crewing URL + B.UUID + '/appraisals'", "seafarers (B)", "CONCAT crewing URL + B.id + '/appraisals'", "seafarers (B)", "Same URL pattern; id source changed", "High", ""),
    ("DOCUMENTS_LINK", "DOCUMENTS_LINK", "CONCAT crewing URL + B.UUID + '/documents'", "seafarers (B)", "CONCAT crewing URL + B.id + '/documents'", "seafarers (B)", "Same URL pattern", "High", ""),
    ("SEA_EXPERIENCE_LINK", "SEA_EXPERIENCE_LINK", "CONCAT crewing URL + B.UUID + '/sea-experience'", "seafarers (B)", "CONCAT crewing URL + B.id + '/sea-experience'", "seafarers (B)", "Same URL pattern", "High", ""),
    ("SEAFARER_PROFILE_LINK", "SEAFARER_PROFILE_LINK", "CONCAT crewing URL + B.UUID + '/personal'", "seafarers (B)", "CONCAT crewing URL + B.id + '/personal'", "seafarers (B)", "Same URL pattern", "High", ""),
    ("CURRENT_RANK_NAME", "CURRENT_RANK_NAME", "A.NAME", "ranks (A) ON A.ID = B.RANK_ID", "A.name", "ranks (A) ON A.id = B.rank_id", "Rank name via FK join", "High", "SAC: landing_zone.db_sac_prod_master_public.ranks; SMAC: curated_db.db_smac_prod_navitasai_masters_public.ranks"),
    ("NATIONALITY_NAME", "NATIONALITY_NAME", "C.NAME", "nationalities (C) ON C.ID = B.NATIONALITY_ID", "C.name", "nationalities (C) ON C.id = B.nationality_id", "Nationality name via FK join", "High", ""),
    ("Last DOC/Contract Company", "Last DOC/Contract Company", "D.NAME", "ship_management_companies (D) ON D.ID = B.CURRENT_COMPANY_ID", "D.name", "companies (D) ON D.id = B.present_doc_company_id", "Present DOC company name", "Medium", "SAC CURRENT_COMPANY_ID → ship_management_companies; SMAC present_doc_company_id → companies. Verify semantic equivalence."),
    ("CONTACT_NUMBER", "CONTACT_NUMBER", "CONCAT(E.COUNTRY_CODE, E.PHONE)", "contact_details (E) ON B.ID = E.SEAFARER_ID", "B.phone", "seafarers (B)", "SAC from contact_details; SMAC direct on seafarers", "Medium", "SMAC no longer concatenates country code. Format may differ."),
    ("EMERGENCY_CONTACT_NUMBER_", "EMERGENCY_CONTACT_NUMBER", "CONCAT(COUNTRY_CODE, LEAD(PHONE) OVER ...)", "contact_details (E)", "CAST(NULL AS STRING)", "N/A", "SAC derived from contact_details window; SMAC explicitly NULL", "Low", "Column renamed (trailing underscore removed). No SMAC equivalent — always NULL in SMAC."),
    ("EMAIL_ID", "EMAIL_ID", "E.EMAIL", "contact_details (E)", "B.email", "seafarers (B)", "SAC from contact row; SMAC on seafarers", "Medium", "SAC may pick email from contact_details row; SMAC uses seafarers.email directly."),
    ("ADDRESS_TYPE", "ADDRESS_TYPE", "CASE on E.CONTACT_TYPE / DENSE_RANK window → PERMANENT/ALTERNATIVE ADDRESS", "contact_details (E)", "'PERMANENT ADDRESS' (hardcoded)", "Derived constant", "SAC dynamic from contact type; SMAC always PERMANENT ADDRESS", "Medium", "SAC filters WHERE NEW_CONTACT_TYPE='1' at end; SMAC hardcodes both ADDRESS_TYPE and NEW_CONTACT_TYPE='1'."),
    ("NEW_CONTACT_TYPE", "NEW_CONTACT_TYPE", "CASE on E.CONTACT_TYPE / ADDRESS_TYPE", "contact_details (E)", "'1' (hardcoded)", "Derived constant", "SAC filters to contact_type=1 rows only", "Medium", "SAC inner filter WHERE C1.NEW_CONTACT_TYPE='1' eliminates alt-address rows; SMAC assumes permanent only."),
    ("PRIMARY_ADDRESS", "PRIMARY_ADDRESS", "E.ADDRESS", "contact_details (E)", "get_json_object(SP.primary_address, '$.address')", "seafarer_profile (SP)", "SAC flat column; SMAC JSON extraction", "Medium", "JSON path $.address in SMAC primary_address field."),
    ("CITY", "CITY", "E.CITY", "contact_details (E)", "get_json_object(SP.primary_address, '$.city')", "seafarer_profile (SP)", "SAC flat; SMAC JSON", "Medium", ""),
    ("PIN_CODE", "PIN_CODE", "E.PIN_CODE", "contact_details (E)", "get_json_object(SP.primary_address, '$.pinCode')", "seafarer_profile (SP)", "SAC flat; SMAC JSON (camelCase pinCode)", "Medium", ""),
    ("NEAREST_AIRPORT", "NEAREST_AIRPORT", "E.NEAREST_AIRPORT", "contact_details (E)", "APT.name", "airports (APT) ON APT.id = get_json_object(SP.primary_address, '$.airportId')", "SAC stored name on contact; SMAC joins airports master", "Medium", "SMAC resolves airportId FK to name."),
    ("STATE", "STATE", "F.NAME", "states (F) ON F.ID = E.STATE_ID", "COALESCE(ST.name, ST_ADDR.name)", "states (ST) ON ST.id = B.state_id; states (ST_ADDR) ON ST_ADDR.id = JSON $.stateId", "SAC from contact state; SMAC seafarer state_id with JSON fallback", "Medium", "SMAC adds fallback from primary_address.stateId when B.state_id NULL."),
    ("COUNTRY", "COUNTRY", "G.NAME", "countries (G) ON G.ID = E.COUNTRY_ID", "CTR.name", "countries (CTR) ON CTR.id = B.country_id", "SAC from contact; SMAC from seafarers.country_id", "Medium", "Source moved from contact_details to seafarers in SMAC."),
    # Sea experience
    ("SEA_EXPERIENCE_ID", "SEA_EXPERIENCE_ID", "CAST(J.ID AS INT)", "sea_experiences (J) ON J.SEAFARER_ID = B.ID", "J.id", "seafarer_sea_experiences (J) ON J.seafarer_id = B.id", "Experience PK; INT→UUID", "High", "Core join: LEFT JOIN all sea experiences per seafarer (not latest only)."),
    ("SIGN_ON_DATE", "SIGN_ON_DATE", "J.FROM_DATE", "sea_experiences (J)", "J.sign_on_date", "seafarer_sea_experiences (J)", "Column rename FROM_DATE→sign_on_date", "High", ""),
    ("SIGN_OFF_DATE", "SIGN_OFF_DATE", "J.TO_DATE", "sea_experiences (J)", "J.sign_off_date", "seafarer_sea_experiences (J)", "Column rename TO_DATE→sign_off_date", "High", ""),
    ("CONTRACT_ID", "CONTRACT_ID", "CAST(J.CONTRACT_ID AS INT)", "sea_experiences (J)", "COALESCE(J.contract_agreement_id, M_FB.id)", "seafarer_sea_experiences + contract_agreements (primary + fallback)", "UUID; fallback agreement id when FK null", "High", "Exclude from SAC int parity; join on experience keys."),
    ("ACTIVE_CONTRACT", "ACTIVE_CONTRACT", "J.ACTIVE_CONTRACT", "sea_experiences (J)", "J.active_contract", "seafarer_sea_experiences (J)", "Direct column mapping", "High", ""),
    ("SAC_CONTRACT", "SAC_CONTRACT", "J.SAC_CONTRACT", "sea_experiences (J)", "CAST(NULL AS BOOLEAN)", "N/A", "SAC-specific flag; no SMAC equivalent", "N/A", "Always NULL in SMAC by design."),
    ("VERIFIED_BY_ID", "VERIFIED_BY_ID", "J.VERIFIED_BY_ID", "sea_experiences (J)", "J.verified_by_id", "seafarer_sea_experiences (J)", "Direct column; INT→UUID", "High", ""),
    ("IS_VERIFIED", "IS_VERIFIED", "J.IS_VERIFIED", "sea_experiences (J)", "J.is_verified", "seafarer_sea_experiences (J)", "Direct column mapping", "High", ""),
    ("VERIFIED_BY_NAME", "VERIFIED_BY_NAME", "J.VERIFIED_BY_NAME", "sea_experiences (J)", "CASE on VERIFIER.first_name/last_name or 'System'", "users (VERIFIER) ON VERIFIER.id = J.verified_by_id", "SAC denormalized on experience; SMAC built from IDP users", "Medium", "SMAC JOIN: curated_db.db_smac_prod_navitasai_idp_public.users"),
    ("VERIFIED_ON", "VERIFIED_ON", "J.VERIFIED_ON", "sea_experiences (J)", "J.verified_at", "seafarer_sea_experiences (J)", "Column rename VERIFIED_ON→verified_at", "High", ""),
    ("SIGN_OFF_REASON", "SIGN_OFF_REASON", "K.NAME", "sign_off_reasons (K) ON K.ID = J.SIGN_OFF_REASON_ID", "K.name", "sign_off_reasons (K) ON K.id = J.sign_off_reason_id", "Lookup join", "High", "SAC master_public; SMAC masters_crewing schema."),
    ("RANK_NAME_SE", "RANK_NAME_SE", "L.NAME", "ranks (L) ON L.ID = J.RANK_ID", "L.name", "ranks (L) ON L.id = J.rank_id", "Sea experience rank name", "High", ""),
    ("IMO_NUMBER", "IMO_NUMBER", "CAST(get_json_object(J.VESSEL_INFO,'$.imo_number') AS INT)", "sea_experiences (J) VESSEL_INFO JSON", "J.imo_number", "seafarer_sea_experiences (J)", "SAC from JSON; SMAC flat column", "High", "SMAC no longer requires JSON parse for IMO."),
    ("VESSEL_NAME", "VESSEL_NAME", "COALESCE(get_json_object(J.VESSEL_INFO,'$.vessel_name'), 'Others')", "sea_experiences (J)", "COALESCE(J.vessel_name, 'Others')", "seafarer_sea_experiences (J)", "Same COALESCE logic; flat column in SMAC", "High", ""),
    ("VESSEL_ID", "VESSEL_ID", "get_json_object(J.VESSEL_INFO,'$.vessel_id')", "sea_experiences (J)", "J.vessel_id", "seafarer_sea_experiences (J)", "SAC JSON; SMAC flat UUID", "High", ""),
    ("SHIP_MANAGEMENT_COMPANY_NAME", "SHIP_MANAGEMENT_COMPANY_NAME", "COALESCE(CASE doc_holder null→external else UPPER(company_name) END, 'Other')", "sea_experiences VESSEL_INFO JSON", "COALESCE(CASE J.doc_holder_company_id IS NULL THEN J.external_company_name ELSE UPPER(COMP_J.name) END, 'Other')", "seafarer_sea_experiences (J) + companies (COMP_J)", "Same business rule; SMAC uses doc_holder_company_id FK", "High", "JOIN: companies COMP_J ON COMP_J.id = J.doc_holder_company_id"),
    ("PORT_OF_REGISTRY_NAME", "PORT_OF_REGISTRY_NAME", "get_json_object(J.VESSEL_INFO,'$.port_of_registry_name')", "sea_experiences (J)", "POR.name", "ports (POR) ON POR.id = J.port_of_registry_id", "SAC denormalized in JSON; SMAC port lookup", "High", ""),
    ("SHIP_MANAGEMENT_COMPANY_ID", "SHIP_MANAGEMENT_COMPANY_ID", "CAST(get_json_object(...,'$.ship_management_company_id') AS INT)", "sea_experiences (J)", "J.doc_holder_company_id", "seafarer_sea_experiences (J)", "SAC from JSON; SMAC FK column", "High", ""),
    ("VESSEL_CATEGORY_NAME", "VESSEL_CATEGORY_NAME", "get_json_object(J.VESSEL_INFO,'$.vessel_category_name')", "sea_experiences (J)", "VCAT.name", "categories (VCAT) ON VCAT.id = J.vessel_category_id", "SAC JSON; SMAC categories lookup", "High", "SMAC: curated_db.db_smac_prod_navitasai_masters_vessel.categories"),
    ("CAPACITY", "CAPACITY", "CONCAT(JSON cargo_capacity.capacity, ' ', UPPER(capacity_unit))", "sea_experiences ADDITIONAL_FIELD JSON", "CONCAT(JSON cargo_capacity_info $.capacity, ' ', $.capacity_unit)", "seafarer_sea_experiences (J)", "Same concat; JSON path renamed", "High", "SAC: $.cargo_capacity.* ; SMAC: $.capacity / $.capacity_unit (PascalCase in engine_spec)."),
    ("DWT", "DWT", "CAST(get_json_object(ADDITIONAL_FIELD,'$.dwt') AS INT)", "sea_experiences (J)", "CAST(J.dwt AS INT)", "seafarer_sea_experiences (J)", "SAC JSON; SMAC flat column", "High", ""),
    ("DUAL_FUEL", "DUAL_FUEL", "CASE JSON dual_fuel='true' THEN 'YES' ELSE 'NO'", "sea_experiences ADDITIONAL_FIELD", "CASE JSON DualFuel='true' THEN 'YES' ELSE 'NO'", "seafarer_sea_experiences engine_specifications", "Same logic; SMAC PascalCase $.DualFuel", "High", ""),
    ("MAKE_NAME", "MAKE_NAME", "JSON $.engine_specification.make_name", "sea_experiences (J)", "JSON $.EngineMakeName", "seafarer_sea_experiences (J)", "JSON path case change", "High", ""),
    ("MODEL_NAME", "MODEL_NAME", "JSON $.engine_specification.model_name", "sea_experiences (J)", "JSON $.EngineModelName", "seafarer_sea_experiences (J)", "JSON path case change", "High", ""),
    ("OUTPUT_POWER", "OUTPUT_POWER", "CONCAT output_power + UPPER(unit) from JSON", "sea_experiences (J)", "CONCAT $.OutputPower + UPPER($.OutputPowerUnit)", "seafarer_sea_experiences (J)", "Same concat logic", "High", ""),
    ("GRT", "GRT", "get_json_object(ADDITIONAL_FIELD,'$.grt')", "sea_experiences (J)", "CAST(J.grt AS STRING)", "seafarer_sea_experiences (J)", "SAC JSON string; SMAC flat grt cast to string", "High", ""),
    ("EXPERIENCE_IN_DAYS", "EXPERIENCE_IN_DAYS", "Onboard: datediff(J.FROM_DATE, current_date()); else J.EXPERIENCE_IN_DAYS (neg in SAC)", "sea_experiences (J)", "Onboard: datediff(current_date(), J.sign_on_date); else COALESCE(J.duration_days,0)", "seafarer_sea_experiences (J)", "SMAC fixes datediff direction for onboard; uses duration_days offboard", "Medium", "INTENTIONAL FIX: SAC onboard datediff produces negative values; SMAC corrected per migration skill."),
    ("EXPERIENCE_IN_MONTHS", "EXPERIENCE_IN_MONTHS", "EXPERIENCE_IN_DAYS / 30", "Derived", "CAST(EXPERIENCE_IN_DAYS AS DOUBLE) / 30", "Derived", "Same formula", "Medium", "Values differ when EXPERIENCE_IN_DAYS differs (onboard case)."),
    ("EXPERIENCE_IN_MONTHS_ROUNDOFF", "EXPERIENCE_IN_MONTHS_ROUNDOFF", "ROUND(EXPERIENCE_IN_DAYS / 30)", "Derived", "ROUND(CAST(EXPERIENCE_IN_DAYS AS DOUBLE) / 30)", "Derived", "Same formula", "Medium", ""),
    ("EXPERIENCE_IN_YEAR", "EXPERIENCE_IN_YEAR", "ROUND((EXPERIENCE_IN_DAYS/30)/12, 1)", "Derived", "ROUND((CAST(EXPERIENCE_IN_DAYS AS DOUBLE)/30)/12, 1)", "Derived", "Same formula", "Medium", ""),
    ("IS_SYNERGY_EXPERIANCE", "IS_SYNERGY_EXPERIANCE", "J.IS_SYNERGY_EXPERIANCE", "sea_experiences (J)", "J.is_inhouse_experience", "seafarer_sea_experiences (J)", "Column rename", "High", "Typo 'EXPERIANCE' preserved in both views."),
    ("POD_NAME", "POD_NAME", "POD.POD from fleets×fleets_vessels×vessels on IMO", "fleets/fleets_vessels/vessels", "POD.POD (same subquery pattern)", "fleets/fld_fleet_vessels/vessels", "Same fleet-vessel-IMO join logic", "High", "JOIN: POD.imo_number = J.imo_number (SAC casts from JSON IMO)."),
    ("POD_VESSEL_NAME", "POD_VESSEL_NAME", "POD.vessel_name", "fleets subquery", "POD.vessel_name", "fleets subquery", "Same source", "High", ""),
    ("VESSEL_CODE", "VESSEL_CODE", "VD.vessel_code from vessel_details", "vessel_details (VD) ON VD.id = JSON vessel_id", "VR.code from vessel_revisions (latest per vessel)", "vessel_revisions (VR) QUALIFY ROW_NUMBER=1", "Different source tables for vessel code", "Medium", "SMAC uses vessel_revisions with ROW_NUMBER partition by vessel_id; SAC uses vessel_details. May produce different codes."),
    ("VESSEL_SUB_CATEGORY", "VESSEL_SUB_CATEGORY", "VSC.NAME via vessel_details.vessel_sub_category_id", "vessel_sub_categories (VSC)", "VSC.name via J.vessel_sub_category_id", "sub_categories (VSC)", "SMAC links sub_category directly on sea experience", "Medium", "SAC indirect via vessel_details; SMAC direct FK on experience."),
    # Appraisals (from appraisals_data view)
    ("FROM_DATE", "FROM_DATE", "N.FROM_DATE (appraisals_data)", "reporting_layer.sac_prod_seafarer_public.appraisals_data (N)", "N.FROM_DATE", "reporting_layer.smac_prod.appraisals_data (N)", "From migrated appraisals view", "High", "JOIN: N.SEAFARER_ID=B.id AND to_date(N.FROM_DATE) BETWEEN sign_on AND COALESCE(sign_off, contract_end) AND vessel_name match"),
    ("TO_DATE", "TO_DATE", "N.TO_DATE", "appraisals_data (N)", "N.TO_DATE", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("STATUS", "STATUS", "N.STATUS", "appraisals_data (N)", "N.STATUS", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("NEED_OF_APPRAISAL", "NEED_OF_APPRAISAL", "N.NEED_OF_APPRAISAL", "appraisals_data (N)", "N.NEED_OF_APPRAISAL", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("APPRAISALS_RANK_NAME", "APPRAISALS_RANK_NAME", "N.APPRAISALS_RANK_NAME", "appraisals_data (N)", "N.APPRAISALS_RANK_NAME", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("APPRAISALS_VESSEL_NAME", "APPRAISALS_VESSEL_NAME", "N.APPRAISALS_VESSEL_NAME", "appraisals_data (N)", "N.APPRAISALS_VESSEL_NAME", "appraisals_data (N)", "From appraisals view; used in join condition", "High", "Join: J.vessel_name = N.APPRAISALS_VESSEL_NAME (SAC from JSON vessel_name)."),
    ("APPRAISALS_VESSEL_CATEGORY_NAME", "APPRAISALS_VESSEL_CATEGORY_NAME", "N.APPRAISALS_VESSEL_CATEGORY_NAME", "appraisals_data (N)", "N.APPRAISALS_VESSEL_CATEGORY_NAME", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("APPRAISAL_DATE", "APPRAISAL_DATE", "N.APPRAISAL_DATE", "appraisals_data (N)", "N.APPRAISAL_DATE", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("IS_MANUAL", "IS_MANUAL", "N.IS_MANUAL", "appraisals_data (N)", "N.IS_MANUAL", "appraisals_data (N)", "From appraisals view", "High", ""),
    ("APPRAISAL_STATUS", "APPRAISAL_STATUS", "CASE WHEN N.FROM_DATE IS NOT NULL THEN 'Completed' ELSE 'Not Completed'", "appraisals_data (N)", "Same CASE", "appraisals_data (N)", "Same derived logic", "High", ""),
    # Contract
    ("CONTRACT_END_DATE", "CONTRACT_END_DATE", "M.END_DATE", "vessel_contracts (M) ON M.ID = J.CONTRACT_ID", "COALESCE(M.end_date, M_FB.end_date, SC_FB.end_date)", "contract_agreements (M) + seafarer_contracts (SC_FB) + contract_agreements (M_FB fallback)", "Primary agreement; fallback when contract_agreement_id null", "High", "v2: 100% match on sample crews after fallback join"),
    ("CONTRACT_START_DATE", "CONTRACT_START_DATE", "M.START_DATE", "vessel_contracts (M)", "COALESCE(M.start_date, M_FB.start_date, SC_FB.start_date)", "Same as CONTRACT_END_DATE", "Same fallback path", "High", ""),
    ("TO_PORT_NAME", "TO_PORT_NAME", "COALESCE(O.NAME, 'Unknown')", "ports (O) ON O.ID = J.TO_PORT_ID", "COALESCE(O.name, 'Unknown')", "ports (O) ON O.id = J.sign_off_port_id", "Sign-off port; column rename TO_PORT→sign_off_port", "High", "SAC db_sac_prod_vessel_public.ports; SMAC masters_public.ports"),
    ("FROM_PORT_NAME", "FROM_PORT_NAME", "COALESCE(P.NAME, 'Unknown')", "ports (P) ON P.ID = J.FROM_PORT_ID", "COALESCE(P.name, 'Unknown')", "ports (P) ON P.id = J.sign_on_port_id", "Sign-on port; FROM_PORT→sign_on_port", "High", ""),
    ("TENTITIVE_SIGN_OFF_DATE", "TENTITIVE_SIGN_OFF_DATE", "COALESCE(J.TO_DATE, M.END_DATE)", "sea_experiences + vessel_contracts", "COALESCE(J.sign_off_date, M.end_date, M_FB.end_date, SC_FB.end_date)", "sea_experiences + resolved contract end", "Uses fallback contract end when FK null", "High", "Typo TENTITIVE preserved."),
    ("CONTRACT_STATUS", "CONTRACT_STATUS", "M.STATUS (raw from vessel_contracts)", "vessel_contracts (M)", "CASE: Void/Cancelled/Terminated→Void; Inactive→Closed; Signed+onboard→InForce; end past→Closed; Signed→Signed", "contract_agreements + seafarer_contracts + sea_experiences", "SAC label parity mapping", "High", "v2: 100% match on sample crews."),
    ("SYNERGY_COMPANY", "SYNERGY_COMPANY", "CASE company='Other' or null id THEN 'FALSE' ELSE COALESCE(Q.SYNERGY_COMPANY,'FALSE')", "ship_management_companies (Q) joined by company name", "CASE doc_holder NULL THEN FALSE ELSE COALESCE(COMP_J.is_inhouse_company, FALSE)", "companies (COMP_J)", "SAC string TRUE/FALSE; SMAC boolean", "Medium", "Type change STRING→BOOLEAN in SMAC. C2 subquery casts to STRING for window logic."),
    ("RECRUITMENT_COMPANY", "RECRUITMENT_COMPANY", "Y.NAME", "ship_management_companies (Y) ON Y.ID=B.HOME_COMPANY_ID AND RECRUITMENT_COMPANY='TRUE'", "REC.name", "companies (REC) ON REC.id = B.recruitment_company_id", "Different FK and lookup approach", "Medium", "SAC filters recruitment companies by flag; SMAC uses dedicated recruitment_company_id."),
    ("AGENT_NAME", "AGENT_NAME", "S.NAME", "agents (S) ON S.ID = B.AGENT_ID", "AGT.name", "agents (AGT) ON AGT.id = B.manning_agent_id", "Agent/manning agent name", "High", "Column rename AGENT_ID→manning_agent_id"),
    ("POSITION_NAME", "POSITION_NAME", "Z.NAME", "positions (Z) ON Z.ID = J.POSITION", "Z.name", "positions (Z) ON Z.id = J.position_id", "Position lookup", "High", "SAC J.POSITION; SMAC J.position_id"),
    ("POSITION_RANK_ID", "POSITION_RANK_ID", "Z.RANK_ID", "positions (Z)", "Z.rank_id", "positions (Z)", "Direct from positions", "High", "INT→UUID"),
    # Remarks
    ("DATE_OF_TERMINATION", "DATE_OF_TERMINATION", "get_json_object(profile_remark,'$.date_of_termination')", "SEAFARER_REMARKS (JSON explode)", "R.date_of_action", "seafarer_remarks (R)", "SAC JSON array; SMAC relational column", "High", "Latest remark per seafarer: ROW_NUMBER PARTITION BY seafarer_id ORDER BY updated_at DESC"),
    ("REMARK", "REMARK", "get_json_object(...,'$.remark')", "SEAFARER_REMARKS", "R.remark_text", "seafarer_remarks (R)", "SAC JSON; SMAC remark_text", "High", ""),
    ("REMARK_TYPE", "REMARK_TYPE", "get_json_object(...,'$.remark_type')", "SEAFARER_REMARKS", "PRT.name", "profile_remark_types (PRT) via profile_remark_type_id", "SAC inline JSON; SMAC type lookup", "High", ""),
    ("INACTIVE_TYPE", "INACTIVE_TYPE", "P.NAME from SEAFARER_PROFILE_REMARKS", "SEAFARER_PROFILE_REMARKS (P)", "PRR.name", "profile_remark_reasons (PRR) via profile_remark_reason_id", "SAC remark_identifier FK; SMAC profile_remark_reasons", "High", ""),
    ("UPDATED_AT", "UPDATED_AT", "N.UPDATED_AT from remarks subquery", "SEAFARER_REMARKS", "R.updated_at", "seafarer_remarks (R)", "Latest remark timestamp", "High", ""),
    ("AVAILABILITY_REMARKS", "AVAILABILITY_REMARKS", "AR.NAME", "availability_remarks (AR) ON AR.ID = B.AVAILABILITY_REMARK_ID", "AR.name", "availability_remarks (AR) ON AR.id = B.availability_remark_id", "Lookup join", "High", ""),
    # Derived / Y subquery / calendar
    ("Overdue by / Days left", "Overdue by / Days left", "CONCAT(datediff(CONTRACT_END_DATE, current_date()), ' DAYS')", "Derived from X.CONTRACT_END_DATE", "Same formula on X.CONTRACT_END_DATE", "Derived", "Same logic in SOURCE_TABLE layer", "High", ""),
    ("LATEST_CONTRACT_END_DATE", "LATEST_CONTRACT_END_DATE", "Y.CONTRACT_END_DATE = MAX(C.end_date) per seafarer", "Y subquery: seafarers A + sea_experiences B + vessel_contracts C", "Y: MAX(COALESCE(CA.end_date, CA_FB.end_date, SC.end_date)) with same fallback joins", "Y subquery with contract_agreements + seafarer_contracts fallback", "Same MAX aggregation; v2 fallback path", "High", "JOIN Y ON X.SEAFARER_ID = Y.SEAFARER_ID"),
    ("LATEST_SIGN_OFF_DATE", "LATEST_SIGN_OFF_DATE", "Y.SIGN_OFF_DATE = MAX(B.TO_DATE)", "Y subquery", "Y.SIGN_OFF_DATE = MAX(B.sign_off_date)", "Y subquery", "Same logic", "High", ""),
    ("LATEST_SIGN_ON_DATE", "LATEST_SIGN_ON_DATE", "Y.SIGN_ON_DATE = MAX(B.FROM_DATE)", "Y subquery", "Y.SIGN_ON_DATE = MAX(B.sign_on_date)", "Y subquery", "Same logic", "High", ""),
    ("LATEST_DATE", "LATE (internal)", "Complex CASE on SEAFARER_TYPE, ONBOARD_SAILING_STATUS vs Y dates", "SOURCE_TABLE X LEFT JOIN Y", "Same CASE logic (mirrored)", "SOURCE_TABLE X LEFT JOIN Y", "Identical business rules for latest date resolution", "High", "Used to derive LATEST_DATE_1 and DATE; not in final output directly."),
    ("LATEST_DATE_1", "LATEST_DATE_1", "CASE WHEN LATEST_DATE IS NULL AND LATEST_SIGN_ON_DATE = SIGN_ON_DATE THEN LATEST_SIGN_OFF_DATE", "Derived in outer C1", "Same CASE", "Derived", "Identical logic", "High", ""),
    ("DATE", "DATE", "COALESCE(LATEST_DATE, LATEST_DATE_1)", "Derived", "COALESCE(LATEST_DATE, LATEST_DATE_1)", "Derived", "Identical logic", "High", ""),
    ("MONTHS", "MONTHS", "GM.MONTHS from GEN_MONTH calendar", "GEN_MONTH: add_months from 2000-01-01 to current_date", "GM.MONTHS (same GEN_MONTH CTE)", "GEN_MONTH calendar CTE", "Monthly grain expansion", "High", "JOIN: GM.MONTHS BETWEEN month_start(sign_on) AND month_start(COALESCE(sign_off, current_date))"),
    # C2 window-derived
    ("COMPANY_STATUS", "COMPANY_STATUS", "Window CASE on STATUS, SYNERGY_COMPANY, DOC_COMPANY, RANK → New Hand / Ex Hand", "C2 subquery on aggregated sea experiences", "Same window CASE (uses PS.code as STATUS)", "C2 subquery", "Logic mirrored; STATUS source differs (B.STATE vs PS.code)", "Medium", "JOIN C2 ON C1.SEA_EXPERIENCE_ID = C2.SEA_EXPERIENCE_ID. SAC STATUS='SIGN_ON'; SMAC STATUS='SIGNON' — verify CASE matches."),
    ("SYNERGY_JOINING_DATE", "SYNERGY_JOINING_DATE", "MIN synergy sign_on when RANK=1 and COMPANY_STATUS in (New Hand, Ex Hand)", "C2 window", "Same window logic", "C2", "Identical formula", "Medium", ""),
    ("SECOND_LATEST_RANK", "SECOND_LATEST_RANK", "LAG(POSITION_NAME) when LAG(RANK)=2 and RANK=1", "C2 window", "Same", "C2", "Identical formula", "High", ""),
    ("FIRST_RANK", "FIRST_RANK", "LAST_VALUE(POSITION_NAME) when first experience", "C2 window", "Same", "C2", "Identical formula", "High", ""),
    ("FIRST_COMPANY", "FIRST_COMPANY", "LAST_VALUE(SHIP_MANAGEMENT_COMPANY_NAME) when first experience", "C2 window", "Same", "C2", "Identical formula", "High", ""),
    ("LATEST_COMPANY", "LATEST_COMPANY", "Onboard: NTH_VALUE(company,2); Onleave: FIRST_VALUE(company)", "C2 window", "Same (SMAC inline without LATEST_COMPANY_1 intermediate)", "C2", "Same business logic", "High", ""),
    ("DOC_COMPANY", "(internal C2 only)", "COALESCE(Q.DOC_COMPANY,'FALSE')", "ship_management_companies (Q)", "CASE WHEN DOC_CS.company_id IS NOT NULL THEN 'TRUE' ELSE 'FALSE' END", "company_services (DOC_CS) service_type_id filter", "Different implementation for DOC flag", "Medium", "SMAC uses company_services junction with static UUID; SAC used ship_management_companies.DOC_COMPANY column. Not in final output."),
    # Final classifications
    ("VESSEL_FLEET_TYPE", "VESSEL_FLEET_TYPE", "CASE UPPER(VESSEL_CATEGORY_NAME) IN (...) THEN DRY/WET", "Derived from VESSEL_CATEGORY_NAME", "Same hardcoded category lists", "Derived", "Identical classification lists", "High", ""),
    ("RANK_LEVEL", "RANK_LEVEL", "CASE on CURRENT_RANK_NAME → Cadet/Management/Operational/Support/Trainee", "Derived", "Same rank lists", "Derived", "Identical classification", "High", ""),
    ("Rank Category", "Rank Category", "CASE on CURRENT_RANK_NAME → Top 4 Rank/Officer/Rating", "Derived", "Same rank lists", "Derived", "Identical classification", "High", ""),
]

join_rows = [
    ("Main seafarer grain", "seafarers B", "seafarers B", "WHERE deleted_at IS NULL (+ SAC _fivetran_deleted=false)", "WHERE deleted_at IS NULL", "High", "Root table for all rows"),
    ("Sea experiences (ALL, not latest)", "sea_experiences J ON J.SEAFARER_ID = B.ID", "seafarer_sea_experiences J ON J.seafarer_id = B.id", "LEFT JOIN; one row per experience per seafarer", "Same", "High", "Core grain multiplier"),
    ("Rank (current)", "ranks A ON A.ID = B.RANK_ID", "ranks A ON A.id = B.rank_id", "LEFT JOIN", "Same", "High", ""),
    ("Nationality", "nationalities C ON C.ID = B.NATIONALITY_ID", "nationalities C ON C.id = B.nationality_id", "LEFT JOIN", "Same", "High", ""),
    ("Present DOC company", "ship_management_companies D ON D.ID = B.CURRENT_COMPANY_ID", "companies D ON D.id = B.present_doc_company_id", "LEFT JOIN", "Same intent", "Medium", "Table/column rename — validate company list parity"),
    ("Recruitment company", "ship_management_companies Y ON Y.ID = B.HOME_COMPANY_ID AND RECRUITMENT_COMPANY='TRUE'", "companies REC ON REC.id = B.recruitment_company_id", "LEFT JOIN", "Different lookup", "Medium", ""),
    ("Contact / Address", "contact_details E ON B.ID = E.SEAFARER_ID (+ states F, countries G on E)", "seafarer_profile SP + states ST/ST_ADDR + countries CTR + airports APT", "SAC multi-row contact; SMAC profile JSON + seafarer columns", "Restructured", "Medium", "SAC filters NEW_CONTACT_TYPE='1' before dedup; SMAC hardcodes permanent address"),
    ("Profile state/status", "Inline B.STATE / B.IS_ACTIVE", "profile_states PS + seafarer_profile_statuses PST", "SAC inline; SMAC normalized masters", "Restructured", "High", ""),
    ("Gender", "Inline B.GENDER enum", "genders GEN ON B.gender_id = GEN.id", "LEFT JOIN", "Restructured", "High", ""),
    ("Sign-off reason", "sign_off_reasons K ON K.ID = J.SIGN_OFF_REASON_ID", "sign_off_reasons K ON K.id = J.sign_off_reason_id", "LEFT JOIN", "Same", "High", "Schema: master_public → masters_crewing"),
    ("Rank (experience)", "ranks L ON L.ID = J.RANK_ID", "ranks L ON L.id = J.rank_id", "LEFT JOIN", "Same", "High", ""),
    ("Contract (primary)", "vessel_contracts M ON M.ID = J.CONTRACT_ID", "contract_agreements M ON M.id = J.contract_agreement_id AND M.deleted_at IS NULL", "LEFT JOIN", "Table swap", "High", "Primary FK path when populated"),
    ("Contract (fallback)", "N/A", "seafarer_contracts SC_FB ON seafarer_id + sign_on_date match WHEN contract_agreement_id IS NULL", "LEFT JOIN", "Reporting-layer workaround", "High", "~90% experiences have null agreement FK"),
    ("Contract (fallback agreement)", "N/A", "contract_agreements M_FB ranked ON SC_FB.id (Signed > Approved)", "LEFT JOIN", "Reporting-layer workaround", "High", "Prefer Signed agreement on parent seafarer_contract"),
    ("Position", "positions Z ON Z.ID = J.POSITION", "positions Z ON Z.id = J.position_id", "LEFT JOIN", "Same", "High", ""),
    ("Appraisals", "appraisals_data N ON N.SEAFARER_ID=B.ID AND date range AND vessel_name", "appraisals_data N ON N.SEAFARER_ID=B.id AND date range AND J.vessel_name=N.APPRAISALS_VESSEL_NAME", "LEFT JOIN", "Same join keys (UUID direct in SMAC)", "High", "SMAC view: reporting_layer.smac_prod.appraisals_data"),
    ("Sign-off/on ports", "ports O ON O.ID=J.TO_PORT_ID; ports P ON P.ID=J.FROM_PORT_ID", "ports O ON O.id=J.sign_off_port_id; ports P ON P.id=J.sign_on_port_id", "LEFT JOIN", "Column rename", "High", ""),
    ("Synergy company check", "ship_management_companies Q ON Q.NAME = derived from VESSEL_INFO", "companies COMP_J ON COMP_J.id = J.doc_holder_company_id", "LEFT JOIN", "FK vs name match", "Medium", ""),
    ("Agent", "agents S ON S.ID = B.AGENT_ID", "agents AGT ON AGT.id = B.manning_agent_id", "LEFT JOIN", "Same", "High", ""),
    ("Availability remarks", "availability_remarks AR ON AR.ID = B.AVAILABILITY_REMARK_ID", "availability_remarks AR ON AR.id = B.availability_remark_id", "LEFT JOIN", "Same", "High", ""),
    ("Vessel code", "vessel_details VD ON VD.id = JSON vessel_id", "vessel_revisions VR latest per vessel_id (ROW_NUMBER)", "LEFT JOIN", "Different source", "Medium", ""),
    ("Vessel sub category", "vessel_sub_categories VSC via vessel_details", "sub_categories VSC ON VSC.id = J.vessel_sub_category_id", "LEFT JOIN", "Direct on experience in SMAC", "Medium", ""),
    ("POD (fleet)", "fleets×fleets_vessels×vessels on IMO", "fleets×fld_fleet_vessels×vessels on IMO", "LEFT JOIN POD.imo_number = J.imo_number", "Same pattern", "High", ""),
    ("Remarks (latest)", "SEAFARER_REMARKS JSON explode + SEAFARER_PROFILE_REMARKS", "seafarer_remarks + profile_remark_types + profile_remark_reasons", "LEFT JOIN R ON R.seafarer_id = B.id", "Restructured", "High", "RNUM=1 latest per seafarer"),
    ("Verified by user", "J.VERIFIED_BY_NAME (denormalized)", "users VERIFIER ON VERIFIER.id = J.verified_by_id", "LEFT JOIN", "SMAC adds join", "Medium", ""),
    ("Vessel category", "From VESSEL_INFO JSON", "categories VCAT ON VCAT.id = J.vessel_category_id", "LEFT JOIN", "Normalized in SMAC", "High", ""),
    ("Port of registry", "From VESSEL_INFO JSON", "ports POR ON POR.id = J.port_of_registry_id", "LEFT JOIN", "Normalized in SMAC", "High", ""),
    ("Y subquery (latest dates)", "MAX sign_on/off/contract_end per seafarer via vessel_contracts", "MAX with COALESCE(CA, CA_FB, SC) fallback — same as inner query", "LEFT JOIN Y ON X.SEAFARER_ID = Y.SEAFARER_ID", "Same logic + v2 fallback", "High", ""),
    ("GEN_MONTH calendar", "add_months 2000-01 to current; BETWEEN sign_on month and sign_off month", "Identical GEN_MONTH CTE and BETWEEN join", "LEFT JOIN", "Same", "High", "Creates monthly grain per experience"),
    ("C2 company status", "Aggregated sea experiences + window functions; JOIN ON SEA_EXPERIENCE_ID", "Same structure; DOC via company_services", "LEFT JOIN C2 ON C1.SEA_EXPERIENCE_ID = C2.SEA_EXPERIENCE_ID", "Same join key", "Medium", ""),
    ("Dedup CTE", "ROW_NUMBER PARTITION BY ~70 columns", "ROW_NUMBER PARTITION BY ~22 key columns", "WHERE RANK_=1", "Different partition scope", "Medium", "SMAC narrower dedup may retain fewer duplicate variants than SAC"),
    ("Contact filter", "WHERE C1.NEW_CONTACT_TYPE = '1'", "No explicit filter (hardcoded NEW_CONTACT_TYPE='1')", "Pre-dedup filter", "Equivalent intent", "High", "Prevents duplicate rows from alt contact addresses in SAC"),
]

cols = [
    "SAC revised_base_view Column",
    "SMAC revised_base_view Column",
    "SAC Source Column",
    "SAC Source Table",
    "SMAC Source Column",
    "SMAC Source Table",
    "Logic Behind the Column",
    "Mapping Confidence",
    "Remark",
]
df_cols = pd.DataFrame(rows, columns=cols)

join_cols = [
    "Join / Layer",
    "SAC Join Condition",
    "SMAC Join Condition",
    "Join Type / Notes",
    "SMAC Equivalent",
    "Mapping Confidence",
    "Remark",
]
df_joins = pd.DataFrame(join_rows, columns=join_cols)

summary_rows = [
    ("Total output columns", "119", "119", "Same column list except EMERGENCY_CONTACT_NUMBER_ → EMERGENCY_CONTACT_NUMBER"),
    ("Row count (validated)", "6,838,743", "3,190,731", "46.7% ratio — SAC has deeper historical monthly grain; expected per migration playbook"),
    ("Architecture", "seafarer × all sea_experiences × GEN_MONTH + C2 windows + dedup", "Identical layered architecture", "SMAC is documented as 1:1 rewrite of SAC logic"),
    ("ID types", "INTEGER IDs throughout", "UUID strings", "Cross-system validation must use CREW_CODE not ID"),
    ("Critical table swaps", "vessel_contracts, contact_details, SEAFARER_REMARKS JSON", "contract_agreements, seafarer_profile JSON, seafarer_remarks relational", "See Join Conditions sheet"),
    ("Known intentional fixes", "EXPERIENCE_IN_DAYS datediff direction (negative onboard)", "datediff(current_date(), sign_on_date) for onboard", "SMAC corrects SAC bug per migration skill"),
    ("Columns always NULL in SMAC", "N/A", "SAC_CONTRACT, EMERGENCY_CONTACT_NUMBER", "No SMAC source equivalent"),
    ("Dedup difference", "Partitions on ~70 columns", "Partitions on ~22 columns", "May explain row count gap alongside SMAC history depth"),
]
df_summary = pd.DataFrame(summary_rows, columns=["Aspect", "SAC", "SMAC", "Notes"])

md_lines = [
    "# revised_base_view — SAC vs SMAC Column Mapping",
    "",
    "> **Validation:** Before flagging mismatches, read [`revised_base_view_validation_notes.md`](./revised_base_view_validation_notes.md) — expected gaps, SAC bugs SMAC fixes, exclusion list for downstream (`planner_view` `REVISED_*`).",
    "",
    "Migration comparison for `reporting_layer.sac_prod_seafarer_public.revised_base_view` → `reporting_layer.smac_prod.revised_base_view`.",
    "",
    "**Source notebooks:**",
    "- SAC: `(Clone) SAC TABLES SET -1.ipynb`",
    "- SMAC: `(Clone) SMAC Reporting Layer Migration SET 1.ipynb`",
    "",
    "## Validation Summary",
    "",
    dataframe_to_md_table(df_summary),
    "",
    "## Architecture (Both Notebooks)",
    "",
    "```",
    "seafarers (B)",
    "  └─ LEFT JOIN all sea_experiences (J)          ← not latest only",
    "       └─ SOURCE_TABLE + Y subquery (latest dates)",
    "            └─ GEN_MONTH calendar join          ← monthly grain",
    "                 └─ C2 window (COMPANY_STATUS, ranks, companies)",
    "                      └─ Dedup CTE (ROW_NUMBER)",
    "                           └─ Final DISTINCT",
    "```",
    "",
    "### Mapping Confidence Breakdown",
    "",
    "| Confidence | Count | Meaning |",
    "| --- | --- | --- |",
    f"| High | {sum(1 for r in rows if r[7] == 'High')} | Direct or equivalent logic |",
    f"| Medium | {sum(1 for r in rows if r[7] == 'Medium')} | Restructured source or semantic/type differences |",
    f"| Low | {sum(1 for r in rows if r[7] == 'Low')} | No SMAC equivalent or partial mapping |",
    f"| N/A | {sum(1 for r in rows if r[7] == 'N/A')} | SAC-specific column with no SMAC source |",
    "",
    "### Key Differences",
    "",
    "1. **Row count gap (expected):** SAC 6,838,743 vs SMAC 3,190,731 (46.7%) — SAC has deeper historical monthly grain.",
    "2. **ID type change:** SAC INT → SMAC UUID. Validate on `CREW_CODE`, not ID.",
    "3. **Intentional fix — `EXPERIENCE_IN_DAYS`:** SMAC corrects SAC datediff direction for onboard seafarers.",
    "4. **Contract join v2:** Primary `contract_agreements`; fallback `seafarer_contracts` + ranked agreement when `contract_agreement_id` null. `CONTRACT_STATUS` maps to SAC labels.",
    "5. **Always NULL in SMAC:** `SAC_CONTRACT`, `EMERGENCY_CONTACT_NUMBER`.",
    "6. **Dedup CTE:** SAC partitions on ~70 columns; SMAC on ~22 key columns.",
    "",
    "## Join Conditions",
    "",
    dataframe_to_md_table(df_joins),
    "",
    "## Column Mapping",
    "",
    dataframe_to_md_table(df_cols),
    "",
]

with open(OUTPUT_MD, "w", encoding="utf-8") as f:
    f.write("\n".join(md_lines))
print(f"Created {OUTPUT_MD}")

try:
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        df_cols.to_excel(writer, sheet_name="Column Mapping", index=False)
        df_joins.to_excel(writer, sheet_name="Join Conditions", index=False)
        df_summary.to_excel(writer, sheet_name="Validation Summary", index=False)

    wb = load_workbook(OUTPUT)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 60)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT)
    print(f"Created {OUTPUT} with {len(df_cols)} column mappings")
except PermissionError:
    print(f"Skipped {OUTPUT} (file may be open in Excel)")
